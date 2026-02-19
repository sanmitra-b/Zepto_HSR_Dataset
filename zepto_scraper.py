"""
Zepto Bangalore Dark Store Product Scraper
==========================================
Endpoint : POST https://bff-gateway.zepto.com/user-search-service/api/v3/search
Strategy : Search alphabet combinations (a-z, aa-az …) to surface all products.
           Each query returns up to 30 products; paginate with pageNumber until
           hasReachedEnd == true or totalProductCount is exhausted.

Setup:
    pip install requests pandas openpyxl

Usage:
    1. Copy the headers below from Chrome DevTools (see SETUP INSTRUCTIONS).
    2. Add store IDs for every Bangalore dark store you want.
    3. Run:  python zepto_scraper.py
"""

import time
import random
import logging
import urllib3

# Disable SSL warnings when verify=False is used
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
import uuid
import requests
import pandas as pd
from datetime import datetime
from itertools import product as iterproduct

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# =============================================================================
# SETUP INSTRUCTIONS
# =============================================================================
# 1. Open https://www.zepto.com in Chrome / Firefox
# 2. DevTools → Network → Fetch/XHR
# 3. Search for anything → find the POST to:
#      bff-gateway.zepto.com/user-search-service/api/v3/search
# 4. Right-click → Copy → Copy as cURL
# 5. Extract these header values and paste below:
# =============================================================================

HEADERS = {
    "accept": "application/json, text/plain, */*",
    "accept-language": "en-US,en;q=0.9",
    "content-type": "application/json",
    "origin": "https://www.zepto.com",
    "referer": "https://www.zepto.com/",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:147.0) Gecko/20100101 Firefox/147.0",

    # ── Copy these EXACTLY from DevTools ──────────────────────────────────────
    "x-xsrf-token":      "ybL14-xcdH-cV-9gZIhUO:51ynKlsYpfGjRB-F8T8fp7VILSs.SA9u0X6U+N1mGgmeMsa6lbsDO9fjvs0S1SqIUtuAU2U",   # ← from request headers
    "x-csrf-secret":     "ieY5UtA9L08",   # ← from request headers
    "request-signature": "b3c9849ae0a76523538c0ef76a1afc644f77e3c26e65b8d1854e06c947d33000",   # ← from request headers (may rotate)
    "x-timezone":        "dffc06574285cdeded5d843257c61fd81de0da07cf397c96694f48b28ede4bd2",    # ← from request headers

    # ── Session / device IDs (copy from your captured request) ────────────────
    "deviceId":   "9b5c4612-a01c-4e12-be2f-dad1a95cd25d",
    "device_id":  "9b5c4612-a01c-4e12-be2f-dad1a95cd25d",
    "sessionId":  "e4cfeeb1-4c1f-49ef-8827-f4590b0317d1",
    "session_id": "e4cfeeb1-4c1f-49ef-8827-f4590b0317d1",

    # ── These are static / can keep as-is ─────────────────────────────────────
    "appVersion":        "14.16.0",
    "app_version":       "14.16.0",
    "platform":          "WEB",
    "tenant":            "ZEPTO",
    "app_sub_platform":  "WEB",
    "marketplace_type":  "SUPER_SAVER",
    "source":            "DIRECT",
    "auth_revamp_flow":  "v2",
    "X-WITHOUT-BEARER":  "true",
    "auth_from_cookie":  "true",

    # NOTE: compatible_components header causes API to return 0 products - removed
}

# =============================================================================
# STORE CONFIGURATION
# Update storeId/store_id/store_ids headers per store before each request,
# OR maintain a separate session per store.
# =============================================================================

STORES = [
    # Working store from your browser session (tested and confirmed)
    {"store_id": "7e5a1821-59ed-4d8a-8431-a3705afb22d2", "store_name": "Active-Store"},
]

# =============================================================================
# SEARCH STRATEGY
# The API doesn't have a "browse all" endpoint — it's query-based.
# API rejects single letters, so we search common products and categories.
# Each unique productVariant.id is deduplicated per store.
# =============================================================================

# Common product categories and items that cover most of the inventory
SEARCH_QUERIES = [
    # Dairy & Breakfast
    "milk", "bread", "butter", "cheese", "paneer", "curd", "yogurt",
    "eggs", "ghee", "cream", "dahi",
    
    #Beverages
    "tea", "coffee", "juice", "water", "cold drink", "coke", "pepsi",
    "energy drink", "soda", "milk shake",
    
    # Snacks & Packaged Foods
    "biscuit", "chips", "namkeen", "chocolate", "candy", "cookies",
    "noodles", "pasta", "maggi", "sauce", "ketchup", "mayo",
    
    # Staples & Cooking
    "rice", "atta", "flour", "dal", "pulses", "oil", "salt", "sugar",
    "spices", "masala", "pickle", "chutney", "papad",
    
    # Fresh Produce
    "vegetables", "fruits", "potato", "onion", "tomato", "banana",
    "apple", "orange", "leafy", "carrot",
    
    # Personal Care
    "shampoo", "soap", "toothpaste", "facewash", "detergent",
    "handwash", "sanitizer", "tissue", "napkin",
    
    # Household
    "cleaner", "dishwash", "floor cleaner", "toilet cleaner",
    "mosquito", "incense", "candle",
    
    # Baby Care
    "diaper", "baby food", "baby care", "wipes",
    
    # Health & Wellness
    "medicine", "vitamins", "protein", "supplements",
    
    # Frozen & Ice Cream
    "ice cream", "frozen food", "frozen vegetables",
    
    # Bakery & Sweets
    "cake", "pastry", "sweets", "mithai", "cookies",
    
    # Meat & Seafood (if available)
    "chicken", "fish", "meat", "seafood", "prawns",
    
    # Pet Supplies
    "pet food", "dog food", "cat food",
    
    # Pharmacy (if applicable)
    "paracetamol", "bandage", "thermometer",
    
    # Generic broad searches
    "organic", "healthy", "fresh", "instant", "ready to eat",
]


PAGE_SIZE      = 30   # API returns max 30 per page (confirmed from response)
SLEEP_MIN      = 1.5  # seconds between requests — be polite
SLEEP_MAX      = 3.0
MAX_PAGES      = 50   # safety cap (30 × 50 = 1500 products per query)

API_URL = "https://bff-gateway.zepto.com/user-search-service/api/v3/search"

# =============================================================================
# PARSER  — maps the exact JSON schema from the real API response
# =============================================================================

def parse_widgets(layout: list, store_name: str) -> list[dict]:
    """
    Extract products from the layout array.
    Products live in widgets with type 'product_grid':
      layout[n].data.resolver.data.items[m].productResponse
    """
    rows = []
    seen_variant_ids = set()
    
    if not layout:
        return rows

    for widget in layout:
        try:
            items = (
                widget
                .get("data", {})
                .get("resolver", {})
                .get("data", {})
                .get("items", []) or []
            )
        except Exception:
            continue

        for item in items:
            pr = item.get("productResponse") or item  # some widgets wrap differently

            product = pr.get("product", {}) or {}
            pv      = pr.get("productVariant", {}) or {}

            variant_id = pv.get("id", "")
            if not variant_id or variant_id in seen_variant_ids:
                continue
            seen_variant_ids.add(variant_id)

            # Images
            images = pv.get("images") or []
            image_path = images[0].get("path", "") if images else ""

            # MRP is stored in paise (1/100 rupee) — divide by 100
            mrp_paise = pv.get("mrp") or pr.get("mrp") or 0
            sp_paise  = pr.get("discountedSellingPrice") or pr.get("sellingPrice") or mrp_paise
            mrp = round(mrp_paise / 100, 2)
            sp  = round(sp_paise  / 100, 2)

            rating_summary = pv.get("ratingSummary") or {}

            rows.append({
                "Store Name":               store_name,
                "Product Name":             product.get("name", ""),
                "Product ID":               product.get("id", ""),
                "Product Variant ID":       variant_id,
                "Brand":                    product.get("brand", ""),
                "MRP (₹)":                  mrp,
                "Selling Price (₹)":        sp,
                "Discount %":               pr.get("discountPercent", 0),
                "Discount Amount (₹)":      round(pr.get("discountAmount", 0) / 100, 2),
                "Available Quantity":       pr.get("availableQuantity", 0),
                "Out of Stock":             pr.get("outOfStock", False),
                "Average Rating":           rating_summary.get("averageRating", ""),
                "Total Ratings":            rating_summary.get("totalRatings", ""),
                "Pack Size":                pv.get("formattedPacksize", ""),
                "Weight (g)":               pv.get("weightInGms", ""),
                "Unit of Measure":          pv.get("unitOfMeasure", ""),
                "Primary Category":         pr.get("primaryCategoryName", ""),
                "L3 Category":              (item.get("l3_details") or {}).get("name", ""),
                "Description":              (product.get("description") or [""])[0],
                "Is Cafe":                  pr.get("isCafe", False),
                "Country of Origin":        product.get("countryOfOrigin", ""),
                "Manufacturer":             product.get("manufacturerName", ""),
                "Image URL":                image_path,
                "Max Allowed Qty":          pv.get("maxAllowedQuantity", ""),
                "Is Sponsored":             any(
                    t.get("type") == "SPONSORED"
                    for t in (pr.get("meta", {}).get("tags") or [])
                ),
            })

    return rows


# =============================================================================
# FETCHER
# =============================================================================

def fetch_query(session: requests.Session, store: dict, query: str, page: int) -> dict | None:
    store_id = store["store_id"]

    # Per-request headers that include store ID
    extra_headers = {
        "storeId":    store_id,
        "store_id":   store_id,
        "store_ids":  store_id,
        "store_etas": f'{{"{store_id}":9}}',
        "requestId":  str(uuid.uuid4()),
        "request_id": str(uuid.uuid4()),
    }

    payload = {
        "query":         query,
        "pageNumber":    page,
        "intentId":      str(uuid.uuid4()),
        "mode":          "TYPED",
    }

    try:
        resp = session.post(
            API_URL,
            headers={**HEADERS, **extra_headers},
            json=payload,
            timeout=20,
            verify=False,
        )
        resp.raise_for_status()
        return resp.json()
    except requests.HTTPError as e:
        log.warning("HTTP %s | store=%s query='%s' page=%d | %s",
                    resp.status_code, store["store_name"], query, page, e)
        # Show response body for debugging
        try:
            error_body = resp.json()
            log.error("API Response: %s", error_body)
        except:
            error_text = resp.text[:500]
            if error_text:
                log.error("API Response: %s", error_text)
        
        if resp.status_code in (401, 403):
            log.error("Auth error — refresh your x-xsrf-token and request-signature from DevTools.")
        return None
    except Exception as e:
        log.error("Request failed | store=%s query='%s' page=%d | %s",
                  store["store_name"], query, page, e)
        return None


def scrape_store(session: requests.Session, store: dict) -> list[dict]:
    store_name = store["store_name"]
    all_rows   = []
    seen_ids   = set()

    log.info("━━ Store: %-20s (%s)", store_name, store["store_id"])

    for query in SEARCH_QUERIES:
        page         = 0
        query_total  = None

        while page <= MAX_PAGES:
            data = fetch_query(session, store, query, page)
            if not data:
                break

            # Pagination metadata
            total_count    = data.get("totalProductCount", 0)
            has_reached_end = data.get("hasReachedEnd", True)
            page_count     = data.get("pageProductCount", 0)

            if query_total is None:
                query_total = total_count

            # Parse products from layout widgets
            layout = data.get("layout", [])
            rows   = parse_widgets(layout, store_name)

            # Deduplicate across queries for this store
            new_rows = []
            for row in rows:
                vid = row["Product Variant ID"]
                if vid not in seen_ids:
                    seen_ids.add(vid)
                    new_rows.append(row)

            all_rows.extend(new_rows)

            log.info("  query='%-4s' page=%2d | page_products=%2d new=%2d "
                     "total_this_store=%d",
                     query, page, page_count, len(new_rows), len(all_rows))

            if has_reached_end or page_count == 0:
                break

            page += 1
            time.sleep(random.uniform(SLEEP_MIN, SLEEP_MAX))

        time.sleep(random.uniform(SLEEP_MIN / 2, SLEEP_MIN))

    log.info("  ✓ Store complete — %d unique products", len(all_rows))
    return all_rows


# =============================================================================
# OUTPUT
# =============================================================================

OUTPUT_COLS = [
    "Store Name", "Product Name", "Product ID", "Product Variant ID",
    "Brand", "MRP (₹)", "Selling Price (₹)", "Discount %",
    "Discount Amount (₹)", "Available Quantity", "Out of Stock",
    "Average Rating", "Total Ratings", "Pack Size", "Weight (g)",
    "Unit of Measure", "Primary Category", "L3 Category", "Description",
    "Is Cafe", "Is Sponsored", "Country of Origin", "Manufacturer",
    "Image URL", "Max Allowed Qty",
]


def write_excel(all_rows: list[dict], filename: str):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    df = pd.DataFrame(all_rows, columns=OUTPUT_COLS)

    HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    CELL_FONT   = Font(name="Arial", size=9)
    ALT_FILL    = PatternFill("solid", fgColor="F5F5F5")
    WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
    THIN        = Border(
        bottom=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin",  color="DDDDDD"),
    )

    def style_sheet(ws, df_sheet):
        ws.freeze_panes = "C2"
        ws.append(list(df_sheet.columns))
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN
        ws.row_dimensions[1].height = 28

        for i, row in enumerate(df_sheet.itertuples(index=False), start=2):
            ws.append(list(row))
            fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
            for cell in ws[i]:
                cell.font  = CELL_FONT
                cell.fill  = fill
                cell.alignment = Alignment(vertical="center")
                cell.border = THIN
            ws.row_dimensions[i].height = 16

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 55)

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="All Products", index=False)
        style_sheet(writer.sheets["All Products"], df)

        for store in STORES:
            name = store["store_name"]
            df_s = df[df["Store Name"] == name].drop(columns=["Store Name"])
            if df_s.empty:
                continue
            sname = name[:31]
            df_s.to_excel(writer, sheet_name=sname, index=False)
            style_sheet(writer.sheets[sname], df_s)

        # Summary sheet
        ws_sum = writer.book.create_sheet("Summary", 0)
        ws_sum.sheet_view.showGridLines = False
        
        # Convert numeric columns to proper types before aggregation
        df_agg = df.copy()
        df_agg["MRP (₹)"] = pd.to_numeric(df_agg["MRP (₹)"], errors='coerce')
        df_agg["Discount %"] = pd.to_numeric(df_agg["Discount %"], errors='coerce')
        df_agg["Average Rating"] = pd.to_numeric(df_agg["Average Rating"], errors='coerce')
        
        summary = (
            df_agg.groupby("Store Name")
            .agg(
                Total_Products   =("Product Variant ID", "count"),
                Unique_Products  =("Product ID",         "nunique"),
                In_Stock         =("Out of Stock",       lambda x: (~x).sum()),
                Out_of_Stock     =("Out of Stock",       "sum"),
                Avg_MRP          =("MRP (₹)",            "mean"),
                Avg_Discount_Pct =("Discount %",         "mean"),
                Avg_Rating       =("Average Rating",     "mean"),
            )
            .reset_index()
        )
        for col in ["Avg_MRP", "Avg_Discount_Pct", "Avg_Rating"]:
            summary[col] = summary[col].round(2)
        style_sheet(ws_sum, summary)

    log.info("Saved %d rows across %d stores → %s",
             len(df), df["Store Name"].nunique(), filename)


# =============================================================================
# MAIN
# =============================================================================

def main():
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M")
    output_file = f"zepto_bangalore_all_stores_{timestamp}.xlsx"

    session = requests.Session()

    all_rows = []
    for store in STORES:
        rows = scrape_store(session, store)
        all_rows.extend(rows)
        time.sleep(random.uniform(2, 4))  # pause between stores

    if all_rows:
        write_excel(all_rows, output_file)
        print(f"\n✅ Done! Output: {output_file}")
        print(f"   Total rows : {len(all_rows)}")
        print(f"   Stores     : {len(STORES)}")
    else:
        log.warning("No data collected. Check your headers (x-xsrf-token, request-signature).")
        log.warning("Re-copy them from Chrome DevTools → Network → the search POST request.")


if __name__ == "__main__":
    main()
